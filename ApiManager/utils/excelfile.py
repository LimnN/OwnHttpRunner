#!/usr/bin/env python
# _*_ coding:utf-8 _*_
import logging

from openpyxl.utils import get_column_letter

from ApiManager.utils.utils import get_xmind_testcase_list, get_absolute_path
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle


# noinspection PyShadowingNames
def xmind2xlsx(xmind_file):
    """Convert XMind file to a xlsx file"""
    xmind_file = get_absolute_path(xmind_file)
    logging.info('Start converting XMind file(%s) to zentao file...', xmind_file)
    testcases = get_xmind_testcase_list(xmind_file)
    wb = Workbook()
    ws = wb.active
    ws.title = 'testcase sheet'
    # TODO: modify excel style
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = True
    ws.page_setup.fitToWidth = 3

    fileheader = ["产品名称", "所属模块", "功能子模块", "用例标题", "步骤", "预期", "关键词", "优先级", "用例类型", "适用阶段"]
    testcase_rows = []
    for testcase in testcases:
        row = gen_a_testcase_row(testcase)
        testcase_rows.append(row)
    ws.append(fileheader)
    for row in testcase_rows:
        ws.append(row)

    column_widths = []
    for row in ws:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell.value)) > column_widths[i]:
                    column_widths[i] = len(str(cell.value))
            else:
                column_widths += [len(str(cell.value))]

    for i, column_width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = column_width

    style = NamedStyle(name='style')
    style.alignment = Alignment(vertical='center', horizontal='left')

    wb.add_named_style(style)

    excel_file = xmind_file[:-6] + '.xlsx'
    wb.save(excel_file)
    # zentao_file = xmind_file[:-6] + '.csv'
    # if os.path.exists(zentao_file):
    #     logging.info('The zentao csv file already exists, return it directly: %s', zentao_file)
    #     return zentao_file
    #
    # with open(zentao_file, 'w', encoding='utf-8') as f:
    #     writer = csv.writer(f)
    #     writer.writerows(testcase_rows)
    #     logging.info('Convert XMind file(%s) to a zentao csv file(%s) successfully!', xmind_file, zentao_file)

    return excel_file


def gen_a_testcase_row(testcase_dict):
    case_product = testcase_dict['product']
    case_module = gen_case_module(testcase_dict['suite'])
    case_submodule = testcase_dict['module']
    case_title = testcase_dict['name']
    # case_precontion = testcase_dict['preconditions']
    case_step, case_expected_result = gen_case_step_and_expected_result(testcase_dict['steps'])
    case_keyword = '功能测试'
    case_priority = gen_case_priority(testcase_dict['importance'])
    case_type = gen_case_type(testcase_dict['execution_type'])
    case_apply_phase = '迭代测试'
    row = [case_product, case_module, case_submodule, case_title, case_step,
           case_expected_result, case_keyword, case_priority, case_type, case_apply_phase]
    return row


def gen_case_module(module_name):
    if module_name:
        module_name = module_name.replace('（', '(')
        module_name = module_name.replace('）', ')')
    else:
        module_name = '/'
    return module_name


def gen_case_step_and_expected_result(steps):
    case_step = ''
    case_expected_result = ''

    for step_dict in steps:
        case_step += str(step_dict['step_number']) + '. ' + step_dict['actions'].replace('\n', '').strip() + '\n'
        case_expected_result += str(step_dict['step_number']) + '. ' + \
            step_dict['expectedresults'].replace('\n', '').strip() + '\n' \
            if step_dict.get('expectedresults', '') else ''

    return case_step, case_expected_result


def gen_case_priority(priority):
    mapping = {1: '高', 2: '中', 3: '低'}
    if priority in mapping.keys():
        return mapping[priority]
    else:
        return '中'


def gen_case_type(case_type):
    mapping = {1: '手动', 2: '自动'}
    if case_type in mapping.keys():
        return mapping[case_type]
    else:
        return '手动'


if __name__ == '__main__':
    # xmind_file = 'C:\\Users\\admin\Desktop\\6级用例.xmind'
    xmind_file = '../../ExampleFile/6级用例.xmind'
    # xmind_file = 'C:\\Users\\admin\Desktop\\5级.xmind'
    xlsx_file = xmind2xlsx(xmind_file)
    print('Convert the xmind file to a csv file successfully: %s', xlsx_file)
